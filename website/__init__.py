from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import SQLAlchemyError
from os import path
from flask_login import LoginManager
from flask_mail import Mail
from flask_login import current_user

db = SQLAlchemy()
mail = Mail()

DB_NAME = "database.db"

fav_list = [] #for retrieve_recipes


def create_app():
    app = Flask(__name__)
    app.config['SECRET_KEY'] = 'abcdefg'
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_NAME}'
    db.init_app(app)

    # email configuration
    app.config["MAIL_SERVER"] = 'smtp.gmail.com'
    app.config["MAIL_PORT"] = 465
    app.config["MAIL_USERNAME"] = 'kathleenchan.wqq@gmail.com'
    app.config['MAIL_PASSWORD'] = 'yynx wgzp ixay ukqe'
    app.config['MAIL_USE_TLS'] = False
    app.config['MAIL_USE_SSL'] = True

    mail.init_app(app)

    from .views import views
    from .auth import auth
    from .staff_auth import staff_auth
    from .staff_views import staff_views

    app.register_blueprint(views, url_prefix='/')
    app.register_blueprint(auth, url_prefix='/')
    app.register_blueprint(staff_auth, url_prefix='/')
    app.register_blueprint(staff_views, url_prefix='/')

    from .models import User, Staff, Post, Comment, Like, Preference

    create_database(app)

    login_manager = LoginManager()
    login_manager.login_view = 'auth.welcome_page'
    login_manager.init_app(app)

    @login_manager.user_loader
    def load_user(username):
        user = User.query.get(str(username))
        staff = Staff.query.get(str(username))

        if user:
            return user
        elif staff:
            return staff

        return None

    # @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@YC's part@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

    from .YCForms import CreateRecipeForm
    import shelve
    from .Recipe import Recipe

    import urllib.request
    import os
    from werkzeug.utils import secure_filename

    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    # ---------#
    import re
    import openpyxl.utils
    from flask_login import login_required
    from .CheckoutForm import CheckoutForm

    # ---------#

    wb = Workbook()

    try:
        wb = load_workbook('website/DB.xlsx')
        ws = wb['Recipes']
        print(f'found existing workbook')
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = [cell for cell in row]
            Recipe.recipes.append(row_data)


    # if DB doesn't exist, create one
    except (KeyError, IOError):
        print('no workbook')
        wb.save('website/DB.xlsx')
        wb = Workbook()
        ws = wb.active
        print('rename title')
        ws.title = 'Recipes'  # rename sheet 1 to 'Recipes'
        Header = ['ID', 'Title', 'Cuisine', 'Skill', 'Time', 'Instruction', 'Ingredient', 'Alt', 'Optional',
                  'Image']
        ws.append(Header)  # adds the headers to the first row

        print('added headers')
        for cell in ws[1]:  # '1' refers to the first row
            cell.font = Font(bold=True)  # makes the font bold



    @app.route('/createRecipe', methods=['GET', 'POST'])
    @login_required
    def create_recipe():
        create_recipe_form = CreateRecipeForm(request.form)
        if request.method == 'POST' and create_recipe_form.validate():
            # Column to check (e.g., 'A', 'B', etc.)
            column_to_check = 'A'  # Change as needed

            # Get the last row number in the column
            last_row = ws.max_row

            # Access the cell in the last row of the specified column
            last_cell = ws[f'{column_to_check}{last_row}']

            # Get the value of the last cell
            last_cell_value = last_cell.value

            try:
                id = int(last_cell_value)
                id += 1

            except ValueError:
                id = last_row

            # made the order match that of the Recipe class
            recipe = Recipe(id, create_recipe_form.title.data, create_recipe_form.image.data,
                            create_recipe_form.skill.data,
                            create_recipe_form.time.data, create_recipe_form.instruction.data,
                            create_recipe_form.cuisine.data,
                            create_recipe_form.ingredient.data,
                            create_recipe_form.alt.data, create_recipe_form.optional.data)

            product_data = list(vars(recipe).values())
            ws.append(product_data)
            Recipe.recipes.append(product_data)

            wb.save('website/DB.xlsx')
            return redirect(url_for('retrieve_recipes'))
        return render_template('createRecipe.html', form=create_recipe_form, user=current_user, staff=True)

    @app.route('/retrieveRecipes')
    @login_required
    def retrieve_recipes():
        recipes_list = []
        for key in ws.iter_rows(min_row=2, values_only=True):  # start reading from row 2 onwards
            recipe = [cell for cell in key]  # look through every cell
            recipes_list.append(recipe)  # add the data from the cell into a list
        print(recipes_list)
        return render_template('retrieveRecipes.html', count=len(recipes_list), recipes_list=recipes_list, user=current_user, staff=True)

    @app.route('/updateRecipe/<int:id>/', methods=['GET', 'POST'])
    @login_required
    def update_recipe(id):
        update_recipe_form = CreateRecipeForm(request.form)
        # Data to find
        data_to_find = int(id)
        recipedata = []
        # Iterate through a specific column (e.g., column A)
        for cell in ws['A']:  # Adjust the column as needed
            if cell.value == data_to_find:
                row_num = cell.row
                # Iterate through each cell in the found row
                for row_cell in ws[row_num]:
                    recipedata.append(row_cell.value)  # or store the value for further processing

                break  # Exit the loop after processing the row

        # if they filled out the form properly, update the page
        if request.method == 'POST' and update_recipe_form.validate():
            # concatenate all the updated data from the form into a list
            updated_data = [data_to_find, (update_recipe_form.title.data), (update_recipe_form.skill.data),
                            (update_recipe_form.time.data), (update_recipe_form.cuisine.data),
                            (update_recipe_form.instruction.data),
                            (update_recipe_form.ingredient.data), (update_recipe_form.alt.data),
                            (update_recipe_form.optional.data), (update_recipe_form.image.data)]

            # look through the Excel sheet and find the ID of the recipe we are trying to update
            for cell in ws['A']:  # 'A' represents the first column
                if cell.value == data_to_find:
                    # Once found, overwrite data in the row
                    for i, new_value in enumerate(updated_data,
                                                  start=1):  # new_value takes from the parameter 'updated_data'
                        # 'i' increases each time the for loop is ran
                        ws.cell(row=cell.row, column=i,
                                value=new_value)  # this line just tells the program which cells should be updated
                        # it does this through getting which row and column the cell is located at
                    break
            try:
                IngredientWS = wb[str(data_to_find)]
                print(f'found existing worksheet')

            # if DB doesn't exist, create one
            except (IOError, KeyError):
                IngredientWS = wb.create_sheet(str(data_to_find))
                IngredientHeader = ['Quantity', 'Measurement', 'Ingredient']
                IngredientWS.append(IngredientHeader)  # adds the headers to the first row
                for cell in IngredientWS[1]:  # '1' refers to the first row
                    cell.font = Font(bold=True)  # makes the font bold
                print('no worksheet found')
                wb.save('website/DB.xlsx')

            # The text containing various ingredients with quantities and measurements
            text = str(update_recipe_form.ingredient.data)

            # Regex pattern to handle ingredients with optional measurements and exclude bracketed words
            # first line captures digits (including decimals)
            # second line captures if 'g' or 'ml' is present
            # third line discards any brackets
            pattern = r'(\d*\.?\d+)' \
                      r'\s*(g|ml)?' \
                      r'\s*([^,(]+)'

            # Find matches using the pattern
            matches = re.findall(pattern, text)

            # Create a list to hold the split sections
            SplitSections = []

            # Loop through the matches and structure them into Quantity, Measurement, Ingredient
            for match in matches:
                quantity, measurement, ingredient = match
                measurement = measurement if measurement else "N/A"  # Assign "N/A" if measurement is missing
                ingredient = ingredient.strip()  # Trim any whitespace from the ingredient
                SplitSections.append([quantity, measurement, ingredient])

            # Write each nested list in the big list to the rows, starting from the second row
            for row_index, sublist in enumerate(SplitSections, start=2):  # Starting at row 2
                for col_index, item in enumerate(sublist, start=1):  # Starting from the first column
                    try:
                        item = float(item)
                    except ValueError:
                        pass
                    cell_ref = f"{openpyxl.utils.get_column_letter(col_index)}{row_index}"
                    IngredientWS[cell_ref] = item
            wb.save('website/DB.xlsx')

            return redirect(url_for('retrieve_recipes'))
        # on initial load, perform the following code
        else:
            # order of which the recipes are shown in updateRecipes.html
            update_recipe_form.title.data = recipedata[1]
            update_recipe_form.skill.data = recipedata[2].lower()
            update_recipe_form.time.data = recipedata[3]
            update_recipe_form.cuisine.data = recipedata[4]
            update_recipe_form.instruction.data = recipedata[5].split('.')
            update_recipe_form.ingredient.data = recipedata[6].split(', ')
            update_recipe_form.alt.data = recipedata[7].split(', ')
            update_recipe_form.optional.data = recipedata[8].split(', ')
            update_recipe_form.image.data = recipedata[9]
            return render_template('updateRecipe.html', form=update_recipe_form, user=current_user, staff=True)

    @app.route('/deleteRecipe/<int:id>', methods=['POST'])
    def delete_recipe(id):

        data_to_find = int(id)


        for cell in ws['A']:
            if cell.value == data_to_find:
                ws.delete_rows(cell.row, 1)  # Delete the row
                break  # Exit the loop after deleting the row
        wb.save('website/DB.xlsx')

        return redirect(url_for('retrieve_recipes'))

    # UPLOAD IMAGE
    UPLOAD_FOLDER = 'static/uploads/'

    app.secret_key = "secret key"
    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

    ALLOWED_EXTENSIONS = set(['png', 'jpg', 'jpeg'])

    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


    @app.route('/uploadImage', methods=['POST'])
    def upload_image():
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No image selected for uploading')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            # print('upload_image filename: ' + filename)
            flash('Image successfully uploaded and displayed below')
            return render_template('index.html', filename=filename)
        else:
            flash('Allowed image types are - png, jpg, jpeg')
            return redirect(request.url)


    @app.route('/display/<filename>')
    def display_image(filename):
        # print('display_image filename: ' + filename)
        return redirect(url_for('static', filename='uploads/' + filename), code=301)

    @app.route('/Recipes', methods=['GET', 'POST'])
    def recipes():
        # right before the page load, look in the Excel and extract all the data into a list
        if request.method == 'POST':
            # Create an instance of the form
            form = CreateRecipeForm(request.form)

            # Check if the form is valid
            if form.validate():
                # Create a new Recipe instance using the form data
                new_recipe = Recipe(
                    title=form.title.data,
                    skill=form.skill.data,
                    time=form.time.data,
                    cuisine=form.cuisine.data,
                    instruction=form.instruction.data,
                    ingredient=form.ingredient.data,
                    alt=form.alt.data,
                    optional=form.optional.data,
                    image='/static'
                )

                Recipe.recipes.append(new_recipe)

        recipe_list = Recipe.recipes
        return render_template("Recipes.html", recipes=recipe_list)

    @app.route("/filterDifficulty")
    def difficulty():
        sortedlist = []
        collected_data = []
        for row in range(2, ws.max_row + 1):  # Starting from row 2
            RData = {}
            first_column_data = ws.cell(row=row, column=1).value  # Data from the first column
            # Data from the same row, but 3 columns to the right
            fourth_column_data = ws.cell(row=row, column=4).value  # Assuming you need to move 3 cells to the right
            # Add the collected data to the list
            RData['id'] = first_column_data
            RData['Skill'] = fourth_column_data.lower()
            collected_data.append(RData)

        for item in collected_data:
            if item['Skill'] == 'easy':
                sortedlist.append(item)
        for item in collected_data:
            if item['Skill'] == 'intermediate':
                sortedlist.append(item)
        for item in collected_data:
            if item['Skill'] == 'hard':
                sortedlist.append(item)

        # Print or process the collected data
        print(collected_data)  # not sorted yet
        print(sortedlist)

        # now that we know the order, match the order of Recipe.recipes to the order of sortedlist
        # Example lists
        list_a = sortedlist
        list_b = Recipe.recipes

        # Create a dictionary from list_b mapping ids to their respective sublists
        dict_b = {item[0]: item for item in list_b}

        # Reorder list_b to match the order of IDs in list_a
        ordered_list_b = [dict_b[dic['id']] for dic in list_a if dic['id'] in dict_b]

        print(ordered_list_b)

        return render_template("filterDifficulty.html", recipes=ordered_list_b)

    @app.route('/filterFavourites')
    def favourites():
        fav_recipes = [recipe for recipe in Recipe.recipes if str(recipe[0]) in fav_list]
        return render_template('filterFavourites.html', fav_recipes=fav_recipes)
    @app.route('/retrieveFavourites', methods=['GET', 'POST'])
    def retrieve_favourites():
        try:
            wb = load_workbook('website/DB.xlsx')
            ws = wb['Fav']
            print(f'opening')
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = [cell for cell in row]
                Recipe.recipes.append(row_data)

        # if DB doesn't exist, create one
        except (KeyError, IOError):
            print('no workbook')
            wb.save('website/DB.xlsx')
            wb = Workbook()
            ws = wb.active
            print('rename title')
            ws.title = 'Fav'  # rename sheet to 'Fav'
            Head = ['ID', 'Title']
            ws.append(Head)  # adds the headers to the first row

            print('added head')
            for cell in ws[1]:  # '1' refers to the first row
                cell.font = Font(bold=True)  # makes the font bold

            wb.save('website/DB.xlsx')

        if request.method == 'POST':
            liked_recipes = request.args.get('liked_recipes', '')
            liked_recipes_list = liked_recipes.split(',')

            print('Recipe', liked_recipes_list)

            if liked_recipes_list[0] not in fav_list:
                fav_list.append(liked_recipes_list[0])
            else:
                fav_list.remove(liked_recipes_list[0])

            print(fav_list)

        fav_recipes = [recipe for recipe in Recipe.recipes if str(recipe[0]) in fav_list]
        print(fav_recipes)

        return render_template('retrieveFavourites.html', count=len(fav_recipes), fav_list=fav_recipes)

    # Search fn
    @app.route("/search")
    def search():
        q = request.args.get("q")
        print(q)

        if q:
            results = [recipe for recipe in Recipe.recipes if q.lower() in recipe[1].lower()]
        else:
            results = []

        return render_template("search_results.html", results=results)


    @app.route('/view_recipe/<int:recipe_id>')
    def view_recipe(recipe_id):
        recipe_details = None
        for recipe in Recipe.recipes:
            if recipe[0] == recipe_id:
                recipe_details = recipe
                break

        return render_template('view_recipes.html', recipe=recipe_details)

    @app.route('/checkout', methods=['GET', 'POST'])
    def checkout():
        Checkout_Form = CheckoutForm(request.form)

        if request.method == 'GET':
            # Load initial data only on GET request
            session['base_price'] = (request.args.get('price', default=None, type=float)) + 3.00
        else:
            # For POST, use the form data submitted by the user
            session['base_price'] = (request.args.get('price', default=None, type=float)) + 3.00

        base_price = session['base_price']  # Add delivery to the base price
        session['price'] = "{:.2f}".format(base_price)

        if request.method == 'POST' and Checkout_Form.validate():
            session['name'] = request.form['name']
            session['email'] = request.form['email']
            session['address'] = request.form['address']
            session['zip'] = request.form['zip']
            session['country'] = request.form['country']
            session['card_number'] = request.form['card_number']
            session['city'] = request.form['city']
            return redirect(url_for('checkout_confirmation'))
        return render_template('checkout.html', form=Checkout_Form, user=current_user)


    @app.route('/checkout_confirmation')
    def checkout_confirmation():
        price = session.get('price')
        name = session.get('name')
        address = session.get('address')
        zip = session.get('zip')
        items = session.get('items')
        city = session.get('city')
        card_number = session.get('card_number')

        masked_credit_card = '*' * 12 + card_number[-4:]

        # delete their cart on confirmation
        CartWB = load_workbook('website/Cart.xlsx')
        UsersCartWS = CartWB[current_user.username]
        CartWB.remove(UsersCartWS)
        CartWB.save('website/Cart.xlsx')

        return render_template('checkout_confirmation.html', price=price, name=name, address=address, zip=zip, items=items, city=city, card_number=masked_credit_card)

    @app.route("/cart")
    @login_required  # This decorator ensures that only logged-in users can access this route
    def cart():
        price = 0  # Initialize total price
        price_ids = []  # List to store price IDs for Stripe
        items = []  # List to store items in the cart
        quantity = []

        try:
            CartWB = load_workbook('website/Cart.xlsx')
            UsersCartWS = CartWB[current_user.username]
            PriceWS = CartWB['Prices']

            # get data from user's cart and structure it such that it is in a list and a nested dictionary inside
            for row in UsersCartWS.iter_rows(min_row=2, max_col=3, values_only=True):
                rowItems = list(row)
                ingredient_dict = {'quantity': rowItems[1], 'ingredient': rowItems[2]}
                items.append(ingredient_dict)

            # go through every item in the list
            for item in items:
                counter = 0
                # go through every item in the worksheet
                for ingredients in PriceWS.iter_cols(max_row=1, values_only=True):
                    counter += 1
                    string = ingredients[0]
                    # if the ingredient name matches
                    if ((item['ingredient']).lower()).replace(" ", "") == ((string.lower()).replace(" ", "")):

                        # get the column number
                        ingredientColumn = openpyxl.utils.get_column_letter(counter)
                        # get the look at the price in the same column
                        quantityNumber = float(PriceWS[f'{ingredientColumn}{2}'].value)
                        # set default price multiplier to 1
                        n = 1
                        # check if the ingredient amount is >= the number on the price sheet
                        while float(item['quantity']) > quantityNumber:
                            n += 1
                            quantityNumber = float(quantityNumber) * n
                        item['buyingQuantity'] = n
                        ingredientPrice = float(PriceWS[f'{ingredientColumn}{3}'].value) * float(n)

                        item['ingredientPrice'] = ingredientPrice
                        itemTotal = ingredientPrice * n
                        item['itemTotal'] = itemTotal
                        price += itemTotal

                        session['items'] = items

        except KeyError:
            items = None
            price = None

        # Render the cart template with the items, total price, price IDs, and quantities
        return render_template('cart.html', items=items, price=price)

    @app.route("/remove/<ingredient>")
    @login_required
    def remove(ingredient):
        CartWB = load_workbook('website/Cart.xlsx')
        UsersCartWS = CartWB[current_user.username]

        # Define the target ingredient you want to delete
        target_ingredient = ingredient

        # Find the column number for 'Ingredient'
        ingredient_col_index = None
        for col in UsersCartWS[1]:  # Assuming the first row contains headers
            if col.value == 'Ingredient':
                ingredient_col_index = col.column  # Get the column index
                break

        # If 'Ingredient' column is found, search for the target ingredient
        if ingredient_col_index:
            # Iterate over the column to find the cell with the target ingredient
            for cell in UsersCartWS.iter_cols(min_col=ingredient_col_index, max_col=ingredient_col_index, min_row=2,
                                        values_only=True):
                for row_index, cell_value in enumerate(cell, start=2):  # Start at 2 to account for header row
                    if cell_value == target_ingredient:
                        # Delete the row
                        UsersCartWS.delete_rows(row_index)
                        # Save the workbook after deletion
                        CartWB.save('website/Cart.xlsx')
                        print(f"Row with {target_ingredient} has been deleted.")
                        break
        else:
            print("Column with 'Ingredient' not found.")

        # Close the workbook to release it from memory
        wb.close()

        return redirect(url_for('cart'))

    @app.route('/add-to-cart', methods=['POST'])
    def add_to_cart():
        data = request.json
        checked_ingredients = data.get('ingredients', [])
        print(data)
        print(checked_ingredients)

        pattern = r'(\d*\.?\d+)' \
                  r'\s*(g|ml)?' \
                  r'\s*([^,(]+)'
        SplitSections = []
        for text in checked_ingredients:
            matches = re.findall(pattern, text)
            for match in matches:
                quantity, measurement, ingredient = match
                measurement = measurement if measurement else "N/A"  # Assign "N/A" if measurement is missing
                ingredient = ingredient.strip()  # Trim any whitespace from the ingredient
                SplitSections.append([quantity, ingredient])

# ------------------------------does the Workbook exist?-------------------------------------
        try:
            CartWB = load_workbook('website/Cart.xlsx')
            print(f'found existing workbook')
        # if DB doesn't exist, create one
        except IOError:
            print('no workbook')
            CartWB = Workbook()

# ------------------------------does the Workbook exist?-------------------------------------

        # ------------------------------does the Worksheet exist?-------------------------------------
        try:
            UsersCartWS = CartWB[current_user.username]
        except KeyError:
            CartWB.create_sheet(current_user.username)
            UsersCartWS = CartWB[current_user.username]
            Header = ['ID', 'Quantity', 'Ingredient']
            UsersCartWS.append(Header)  # adds the headers to the first row
            for cell in UsersCartWS[1]:  # '1' refers to the first row
                cell.font = Font(bold=True)  # makes the font bold

        # ------------------------------does the Worksheet exist?-------------------------------------

# ----------------------------------add the data to the worksheet of the respective users----------------------------
        max_rows = UsersCartWS.max_row
        # we check if their cart is empty, if it is start writing from row 2 onwards

        # Write each nested list in the big list to the rows, starting from the second row
        for row_index, sublist in enumerate(SplitSections, start=(max_rows + 1)):  # Starting at row 2
            for col_index, item in enumerate(sublist, start=1):  # Starting from the first column
                try:
                    item = float(item)
                    item = item * float(data['servings'])
                except ValueError:
                    pass
                cell_ref = f"{openpyxl.utils.get_column_letter(col_index + 1)}{row_index}"
                add_to_first_col = f"{openpyxl.utils.get_column_letter(1)}{row_index}"
                UsersCartWS[cell_ref] = item
                UsersCartWS[add_to_first_col] = int(data['id'])
        CartWB.save('website/Cart.xlsx')
# ----------------------------------add the data to the worksheet of the respective users----------------------------
        # Respond to the client
        return jsonify({'status': 'success', 'message': 'Ingredients added to cart'})

    # ---------------------WQ's part------------------------#
    import mpld3
    from flask import render_template, request, redirect, url_for, jsonify, session
    from .Forms import CreateUserForm, CreateCustomerForm, CreateFeedbackForm, CreateSurveyForm
    import shelve
    from .Feedback import Feedback
    from .Survey import Survey
    from datetime import datetime
    import os

    # Important need to be included
    app.secret_key = os.urandom(24)  # For session security
    # Start of the Project

    # Built-in function for the template of email like from, to Subject header and body
    from email.message import EmailMessage
    # Add security
    import ssl
    import smtplib

    current_timestamp = datetime.now()
    formatted_timestamp = current_timestamp.strftime("%d/%m/%Y %I:%M%p")
    print(formatted_timestamp)

    @app.route("/FeedbackForm", methods=["GET", "POST"])
    @login_required
    def feedback_form():
        # It helps ensure that the data submitted aligns with the authenticated user's information (readonly)
        # Change the name and email to retrieve from the user that login
        name = current_user.username
        email = current_user.email

        create_feedback_form = CreateFeedbackForm(request.form)
        if request.method == "POST" and create_feedback_form.validate():
            # Dummy Date
            feedback_dict = {
                1: Feedback("16/12/2023 05:56PM", "Test1", "test1123@gmail.com", "test1subject",
                            "test1feedback"),
                2: Feedback("16/12/2023 06:07PM", "Test2", "test2123@gmail.com", "test2subject",
                            "test2feedback")}
            db = shelve.open('feedback.db', "c")
            try:
                feedback_dict = db["Feedback"]
            except:
                print("Error in retrieving feedback from feedback.db")
            feedback = Feedback(formatted_timestamp, create_feedback_form.username.data,
                                create_feedback_form.email_address.data, create_feedback_form.subject.data,
                                create_feedback_form.feedback.data)
            feedback_dict[feedback.get_feedback_id()] = feedback
            db["Feedback"] = feedback_dict
            print(feedback_dict)

            email_password = "wogwjgkvzgliubck"
            username = feedback.get_username()
            email_address = feedback.get_email_address()
            subject = feedback.get_subject()
            feedback = feedback.get_feedback()
            combined_feedback = feedback + "\n\n" + email_address

            Default_message_Subject = "Feedback"
            Default_message_body = "Hi " + username + "\n\n" + "Thank you so much for taking the time to share your feedback! We will reply to any inquires you may have within 2-3 business days" + "\n" + "Thank you again for being a loyal customer" + "\n\n" + "Best Regards" + "\n" + "Eco-Eats Management"

            print(username)
            print(email_address)
            print(subject)
            print(feedback)

            em = EmailMessage()
            em["From"] = "weeqichew0316@gmail"
            em["To"] = "weeqichew0316@gmail.com"  # Email Reciever
            em["Subject"] = subject
            em.set_content(combined_feedback)

            context = ssl.create_default_context()

            # Sending customer feedback to staff gmail
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
                smtp.login("weeqichew0316@gmail.com", email_password)
                smtp.sendmail("weeqichew0316@gmail.com", "weeqichew0316@gmail.com", em.as_string())

            em1 = EmailMessage()
            em1["From"] = "weeqichew0316@gmail"
            em1["To"] = email_address  # Email Reciever
            em1["Subject"] = Default_message_Subject
            em1.set_content(Default_message_body)

            context = ssl.create_default_context()

            # Sending thank you message to customer
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as smtp:
                smtp.login("weeqichew0316@gmail.com", email_password)
                smtp.sendmail("weeqichew0316@gmail.com", email_address, em1.as_string())

            db.close()

            return redirect(url_for("confirmation"))
        return render_template('feedback.html', form=create_feedback_form, name=name, email=email)

    @app.route("/retrieveFeedback")
    @login_required
    def retrieve_feedback():
        feedback_dict = {1: Feedback("16/12/2023 05:56PM", "Test1", "test1123@gmail.com", "test1subject",
                                     "test1feedback"),
                         2: Feedback("16/12/2023 06:07PM", "Test2", "test2123@gmail.com", "test2subject",
                                     "test2feedback")}
        db = shelve.open("feedback.db", "r")
        feedback_dict = db["Feedback"]

        db.close()

        feedback_list = []
        for key in feedback_dict:  # Key has to be a number
            feedback = feedback_dict.get(key)
            feedback_list.append(feedback)

        print(feedback_list)

        return render_template('retrieveFeedback.html', count=len(feedback_list), feedback_list=feedback_list, user=current_user, staff=True)

    @app.route('/deleteFeedback/<int:id>', methods=['POST'])
    def delete_feedback(id):
        db = shelve.open("feedback.db", "w")
        feedback_dict = db["Feedback"]
        print(id)
        print(feedback_dict[id])
        print(feedback_dict[id].get_feedback())
        feedback_dict.pop(id)
        db["Feedback"] = feedback_dict
        db.close()

        return redirect(url_for('retrieve_feedback'))

    from .chat import get_response

    @app.get("/chatbot")
    def index_get():
        return render_template("base.html")

    # chabot
    @app.post("/predict")
    def predict():
        text = request.get_json().get("message")
        # TODO: check if text is valid
        response = get_response(text)
        message = {"answer": response}
        return jsonify(message)

    @app.route("/SurveyForm", methods=["GET", "POST"])
    @login_required
    def survey_form():
        # Change the name and email to retrieve from the user that login
        name = current_user.username
        email = current_user.email
        create_survey_form = CreateSurveyForm(request.form)
        # Limit the user to only submit 1 form
        if session.get("form_submitted"):
            # Change the faq.html to html that says "you have already submitted a form"
            return render_template("already_submit.html")
        if request.method == "POST" and create_survey_form.validate():
            # Dummy Data
            survey_dict = {
                "1": Survey("16/12/2023 12.33PM", "abc123", "abc123@gmail.com", "Chinese", "Orange Chicken"),
                "2": Survey("16/12/2023 12:45PM", "abc234", "abc234@gmail.com", "Korean", "Seaweed Soup"),
                "3": Survey("16/12/2023 02:45PM", "Test1", "Test1@gmail.com", "Indian", "Curry Chicken Masala"),
                "4": Survey("16/12/2023 05.25PM", "Test2", "Test2@gmail.com", "Mexican", "Tacos"),
                "5": Survey("16/12/2023 06:20PM", "Test3", "Test3@gmail.com", "Chinese", "Hor Fun"),
                "6": Survey("16/12/2023: 11:20PM", "Test4", "Test4@gmail.com", "Japanese", "Takoyaki"),
                "7": Survey("18/12/2023 07:20AM", "Test5", "Test5@gmail.com", "Korean", "Kimchi Soup"),
                "8": Survey("18/12/2023: 11:20AM", "Test6", "Test6@gmail.com", "Thai", "Pad Thai"),
                "9": Survey("18/12/2023: 01.20PM", "Test7", "Test7@gmail.com", "Korean", "Bulgogi"),
                "10": Survey("18/12/2023 05:20PM", "Test8", "Test8@gmail.com", "Thai", "Tom Yum Soup")
            }

            db = shelve.open("survey.db", "c")
            try:
                survey_dict = db["Survey"]
            except:
                print("Error in retrieving survey responses from survey.db")
            survey = Survey(formatted_timestamp, create_survey_form.username.data,
                            create_survey_form.email_address.data, create_survey_form.cuisine.data,
                            create_survey_form.recipe_name.data)
            survey.increment(create_survey_form.cuisine.data)
            survey_dict[survey.get_survey_id()] = survey
            db["Survey"] = survey_dict

            print(survey_dict)
            username = survey.get_username()
            email_address = survey.get_email_address()
            cuisine = survey.get_cuisine()
            recipe_name = survey.get_recipe_name()
            print(username)
            print(email_address)
            print(cuisine)
            print(recipe_name)

            db.close()
            session["form_submitted"] = True

            return redirect(url_for("confirmation"))  # The function def confirmation()
        return render_template("survey.html", form=create_survey_form, name=name, email=email)

    @app.route("/retrieveSurvey")
    @login_required
    def retrieve_survey():
        # Dummy Data
        survey_dict = {
            "1": Survey("16/12/2023 12.33PM", "abc123", "abc123@gmail.com", "Chinese", "Orange Chicken"),
            "2": Survey("16/12/2023 12:45PM", "abc234", "abc234@gmail.com", "Korean", "Seaweed Soup"),
            "3": Survey("16/12/2023 02:45PM", "Test1", "Test1@gmail.com", "Indian", "Curry Chicken Masala"),
            "4": Survey("16/12/2023 05.25PM", "Test2", "Test2@gmail.com", "Mexican", "Tacos"),
            "5": Survey("16/12/2023 06:20PM", "Test3", "Test3@gmail.com", "Chinese", "Hor Fun"),
            "6": Survey("16/12/2023: 11:20PM", "Test4", "Test4@gmail.com", "Japanese", "Takoyaki"),
            "7": Survey("18/12/2023 07:20AM", "Test5", "Test5@gmail.com", "Korean", "Kimchi Soup"),
            "8": Survey("18/12/2023: 11:20AM", "Test6", "Test6@gmail.com", "Thai", "Pad Thai"),
            "9": Survey("18/12/2023: 01.20PM", "Test7", "Test7@gmail.com", "Korean", "Bulgogi"),
            "10": Survey("18/12/2023 05:20PM", "Test8", "Test8@gmail.com", "Thai", "Tom Yum Soup")
        }
        db = shelve.open("survey.db", "r")
        survey_dict = db["Survey"]
        db.close()

        survey_list = []
        for key in survey_dict:
            survey = survey_dict.get(key)
            survey_list.append(survey)

        return render_template("retrieveSurvey.html", count=len(survey_list), survey_list=survey_list, user=current_user, staff=True)

    @app.route('/deleteSurvey/<int:id>', methods=['POST'])
    def delete_survey(id):
        db = shelve.open("survey.db", "w")
        survey_dict = db["Survey"]
        print(id)
        print(survey_dict[id])
        print(survey_dict[id].get_cuisine())
        cuisine_delete = survey_dict[id].get_cuisine()

        # Just placeholder so that I can use the instance
        survey = Survey("16/12/2023 12.33PM", "abc123", "abc123@gmail.com", "Chinese", "Orange Chicken")
        survey.decrement(cuisine_delete)
        survey_dict.pop(id)
        db["Survey"] = survey_dict
        db.close()

        return redirect(url_for('retrieve_survey'))

    import matplotlib.pyplot as plt

    @app.route("/surveyChart")
    @login_required
    def display_chart():
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.bar(["Chinese", "Korean", "Japanese", "Mexican", "Thai", "Indian"],
               [Survey.chinese_id, Survey.korean_id, Survey.japanese_id, Survey.mexican_id,
                Survey.thai_id, Survey.indian_id], width=0.5)
        ax.set_xticklabels(["", "Chinese", "", "Korean", "", "Japanese", "", "Mexican", "", "Thai", "", "Indian"])
        chart_html = mpld3.fig_to_html(fig)

        return render_template("surveyChart.html", bar_chart=chart_html, user=current_user, staff=True)

    @app.route("/confirmation")
    def confirmation():
        return render_template("confirmation.html")

    @app.route("/faq")
    def faq():
        return render_template("faq.html")

    return app


def create_database(app):
    try:
        if not path.exists('website/' + DB_NAME):
            with app.app_context():
                db.create_all()
            print('Created Database!')

    except SQLAlchemyError as e:
        print(f"Error creating database: {str(e)}")
